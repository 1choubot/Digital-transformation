# -*- coding: utf-8 -*-
"""
员工周工作对比考核表 — 精简版
仅含"周报 vs 日报逐日对照"核心表格。
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

OUTPUT = "E:/Digital-transformation/docs/员工周工作对比考核表_模板.xlsx"

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "周vs日对照"

# ========== 样式 ==========
thin = Border(left=Side('thin'),right=Side('thin'),top=Side('thin'),bottom=Side('thin'))

def apply_style(cell, font=None, fill=None, alignment=None, border=None, numfmt=None):
    if font:      cell.font = font
    if fill:      cell.fill = fill
    if alignment: cell.alignment = alignment
    if border:    cell.border = border
    if numfmt:    cell.number_format = numfmt

title_font   = Font(name='微软雅黑', size=15, bold=True, color='1F4E79')
hdr_font     = Font(name='微软雅黑', size=10, bold=True, color='FFFFFF')
data_font    = Font(name='微软雅黑', size=10, color='333333')
dim_font     = Font(name='微软雅黑', size=9, color='AAAAAA')

title_fill   = PatternFill('solid', fgColor='D6E4F0')
hdr_fill     = PatternFill('solid', fgColor='2E75B6')
plan_fill    = PatternFill('solid', fgColor='DAEEF3')     # 周报列 — 浅蓝
actual_fill  = PatternFill('solid', fgColor='FFFFFF')     # 日报列 — 白色

center = Alignment(horizontal='center', vertical='center', wrap_text=True)
left   = Alignment(horizontal='left',   vertical='center', wrap_text=True)

# ========== 周报数据（陈芋如 6.15-6.21 工作总结） ==========
# (序号, 工作任务/项目, 工作目标/自述内容, 计划时间, 完成时间)
weekly = [
    (1, 'KRF26022节点仪项目改造',           '实现整线正常运行',                                              '2026-06-15', '2026-06-15'),
    (2, '公司项目经理模式管理群',            '完成"基于数字孪生的设备故障预测与健康管理平台_项目方案"（100%）', '2026-06-15', '2026-06-15'),
    (3, 'KRF26033燃烧窑温度监测项目',        '估计上位机成本，完成成本估算表',                                 '2026-06-16', '2026-06-16'),
    (4, 'KRF202617军团（AI+研发）',          '配合人事部门人员在work Buddy建立筛选简历的自动化任务（70%）',    '2026-06-17', '2026-06-17'),
    (5, 'KRF202617军团（AI+研发）',          '配合人事部门人员在work Buddy建立筛选简历的自动化任务（100%）',   '2026-06-18', '2026-06-17'),
    (6, 'KRF202617军团（AI+研发）',          '实现work Buddy和Qclaw接入钉钉企业机器人',                        '新增',       '2026-06-18'),
    (7, 'KRF202617军团（AI+研发）',          '解决因为权限问题导致Qclaw无法获取群文件的问题',                   '新增',       '2026-06-21'),
]

# 工作日（6.15周一 ~ 6.21周日）
dates = ['2026-06-15','2026-06-16','2026-06-17','2026-06-18','2026-06-19','2026-06-20','2026-06-21']
wdays = {'2026-06-15':'周一','2026-06-16':'周二','2026-06-17':'周三',
         '2026-06-18':'周四','2026-06-19':'周五','2026-06-20':'周六','2026-06-21':'周日'}

# 列宽
widths = [12, 6, 22, 34, 18, 36, 10, 14, 14]  # A-I

# ========== Row 1: 标题 ==========
ws.merge_cells('A1:I1')
c = ws.cell(row=1, column=1, value='周报 vs 日报逐日对照表')
apply_style(c, font=title_font, fill=title_fill, alignment=center)
ws.row_dimensions[1].height = 34

# ========== Row 2: 表头 ==========
headers = ['日期','星期','周报的任务\n（工作总结——工作任务）','周报工作总结\n（自述）',
           '日报的任务\n（项目名称）','日报实际工作内容','完成进度','实际完成时间\n（以日报为准）','周报完成时间']
for i,(h,w) in enumerate(zip(headers, widths), 1):
    c = ws.cell(row=2, column=i, value=h)
    apply_style(c, font=hdr_font, fill=hdr_fill, alignment=center, border=thin)
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
ws.row_dimensions[2].height = 36

# ========== Row 3+: 每天数据 ==========
row = 3
for date_str in dates:
    wd = wdays[date_str]
    # 匹配此日期的周报项（计划时间匹配；若为"新增"则按完成时间匹配）
    matches = [w for w in weekly if w[3] == date_str or (w[3] == '新增' and w[4] == date_str)]

    if matches:
        for seq, proj, goal, plan_d, comp_d in matches:
            r = row
            ws.row_dimensions[r].height = 38 if len(goal) > 40 else 30

            row_data = [
                (date_str, center, plan_fill),
                (wd,       center, plan_fill),
                (proj,     left,   plan_fill),
                (goal,     left,   plan_fill),
                ('（填写日报项目名）', left, actual_fill),
                ('（填写日报实际内容）', left, actual_fill),
                (None,     center, actual_fill),
                ('（填写）', center, actual_fill),
                (comp_d,   center, plan_fill),
            ]
            for ci, (val, align, fill) in enumerate(row_data, 1):
                c = ws.cell(row=r, column=ci, value=val)
                apply_style(c, font=data_font, fill=fill, alignment=align, border=thin)
            ws.cell(row=r, column=7).number_format = '0%'
            row += 1
    else:
        # 无周报项 — 仅日报行
        r = row
        ws.row_dimensions[r].height = 28
        row_data = [
            (date_str, center, plan_fill),
            (wd,       center, plan_fill),
            ('—',      center, plan_fill),
            ('（周报无此日总结项）', left, plan_fill),
            ('（填写日报项目名）',   left, actual_fill),
            ('（填写日报实际内容）', left, actual_fill),
            (None,     center, actual_fill),
            ('（填写）', center, actual_fill),
            ('—',      center, plan_fill),
        ]
        for ci, (val, align, fill) in enumerate(row_data, 1):
            c = ws.cell(row=r, column=ci, value=val)
            apply_style(c, font=data_font, fill=fill, alignment=align, border=thin)
        ws.cell(row=r, column=7).number_format = '0%'
        row += 1

# ========== 数据验证：完成进度 0~1 ==========
dv = DataValidation(type='decimal', operator='between', formula1='0', formula2='1')
dv.error = '请输入0~1之间的小数（0.8=80%）'
ws.add_data_validation(dv)
for rr in range(3, row):
    dv.add(ws.cell(row=rr, column=7))

# ========== 冻结 ==========
ws.freeze_panes = 'A3'

# ========== 打印 ==========
ws.page_setup.orientation = 'landscape'
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0

wb.save(OUTPUT)
print(f'OK: {OUTPUT}')
